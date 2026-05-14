from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Count, Avg, Q
from django.db.models.functions import TruncMonth
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from .forms import UserRegisterForm, UserProfileForm, ManagerStudentForm
from .decorators import role_required
from django.contrib.auth import login
import os
from rest_framework_simplejwt.tokens import RefreshToken

from django.contrib.auth import get_user_model
from urllib.parse import urlencode
from pathlib import Path
import csv
import io
from courses.forms import (
    CourseForm,
    CourseAssignForm,
    TaskForm,
    TaskAssignForm,
    TaskSubmissionForm,
    CourseMaterialForm,
)
from courses.models import Course, Enrollment, Task, TaskAssignment, TaskSubmission, Lesson, LessonAsset
from courses.models import CourseMaterial
from tests.forms import (
    TestForm,
    QuestionForm,
    AnswerFormSet,
    MatchingPairFormSet,
    OrderingItemFormSet,
)
from tests.models import (
    Test,
    Question,
    Answer,
    MatchingPair,
    OrderingItem,
    QUESTION_TYPE_SINGLE,
    QUESTION_TYPE_MULTI,
    QUESTION_TYPE_MATCHING,
    QUESTION_TYPE_ORDERING,
    QUESTION_TYPE_SHORT,
    QUESTION_TYPE_LONG,
    MANUAL_QUESTION_TYPES,
)
from results.models import TestResult, TestAnswer
from analytics.models import TrainingEvent, Dashboard, Report
from users.models import UserCompetency, Specialty, Position
from feedback.models import TaskReview, Feedback

User = get_user_model()


def _analytics_filters_from_payload(payload, base=None):
    base = base or {}
    def _get(key):
        return (payload.get(key) or '').strip()
    base['start'] = _get('start') or base.get('start') or ''
    base['end'] = _get('end') or base.get('end') or ''
    base['department'] = _get('department') or base.get('department') or ''
    base['specialty'] = _get('specialty') or base.get('specialty') or ''
    base['position'] = _get('position') or base.get('position') or ''
    base['course'] = _get('course') or base.get('course') or ''
    base['group_by'] = _get('group_by') or base.get('group_by') or 'department'
    return base


def _analytics_filters_from_request(request):
    base = {}
    dashboard_id = (request.GET.get('dashboard') or '').strip()
    if dashboard_id.isdigit():
        dashboard = (
            Dashboard.objects.filter(
                Q(owner=request.user) | Q(is_shared=True),
                id=int(dashboard_id),
            )
            .only('config')
            .first()
        )
        if dashboard:
            base = dashboard.config or {}
            base['dashboard'] = dashboard_id
    return _analytics_filters_from_payload(request.GET, base=base)


def _apply_date_filter(qs, field, start, end, date_field=False):
    if start:
        key = f"{field}__gte" if date_field else f"{field}__date__gte"
        qs = qs.filter(**{key: start})
    if end:
        key = f"{field}__lte" if date_field else f"{field}__date__lte"
        qs = qs.filter(**{key: end})
    return qs


def _build_analytics_data(filters):
    start = parse_date(filters.get('start') or '') if filters.get('start') else None
    end = parse_date(filters.get('end') or '') if filters.get('end') else None
    department = filters.get('department') or None
    specialty_id = filters.get('specialty') or None
    position_id = filters.get('position') or None
    course_id = filters.get('course') or None

    users_qs = User.objects.filter(role='EMPLOYEE', is_active=True)
    if department:
        users_qs = users_qs.filter(department=department)
    if specialty_id:
        users_qs = users_qs.filter(specialty_id=specialty_id)
    if position_id:
        users_qs = users_qs.filter(position_id=position_id)

    courses_qs = Course.objects.filter(enrollments__user__in=users_qs).distinct()
    if course_id:
        courses_qs = courses_qs.filter(id=course_id)

    enrollments = Enrollment.objects.filter(user__in=users_qs, course__in=courses_qs)
    enrollments = _apply_date_filter(enrollments, 'assigned_at', start, end)

    test_results = TestResult.objects.filter(
        user__in=users_qs,
        test__course__in=courses_qs,
    )
    test_results = _apply_date_filter(test_results, 'completed_at', start, end)

    submissions = TaskSubmission.objects.filter(
        assignment__user__in=users_qs,
        assignment__task__course__in=courses_qs,
    )
    submissions = _apply_date_filter(submissions, 'submitted_at', start, end)

    summary = {
        'employees': users_qs.count(),
        'courses_assigned': enrollments.count(),
        'courses_completed': enrollments.filter(status=Enrollment.Status.COMPLETED).count(),
        'courses_in_progress': enrollments.filter(status=Enrollment.Status.IN_PROGRESS).count(),
        'courses_overdue': enrollments.filter(status=Enrollment.Status.OVERDUE).count(),
        'tests_taken': test_results.count(),
        'avg_score': test_results.aggregate(avg=Avg('score')).get('avg'),
        'pass_rate': 0,
        'tasks_submitted': submissions.count(),
        'tasks_approved': submissions.filter(status=TaskSubmission.Status.APPROVED).count(),
    }
    summary['completion_rate'] = (
        round(summary['courses_completed'] / summary['courses_assigned'] * 100, 1)
        if summary['courses_assigned']
        else 0
    )
    passed_count = test_results.filter(passed=True).count()
    if summary['tests_taken']:
        summary['pass_rate'] = round(passed_count / summary['tests_taken'] * 100, 1)
    summary['task_approval_rate'] = (
        round(summary['tasks_approved'] / summary['tasks_submitted'] * 100, 1)
        if summary['tasks_submitted']
        else 0
    )
    summary['overdue_rate'] = (
        round(summary['courses_overdue'] / summary['courses_assigned'] * 100, 1)
        if summary['courses_assigned']
        else 0
    )

    enrollment_by_user = {
        row['user_id']: row
        for row in enrollments.values('user_id').annotate(
            assigned=Count('id'),
            completed=Count('id', filter=Q(status=Enrollment.Status.COMPLETED)),
            in_progress=Count('id', filter=Q(status=Enrollment.Status.IN_PROGRESS)),
            overdue=Count('id', filter=Q(status=Enrollment.Status.OVERDUE)),
            progress_avg=Avg('progress'),
        )
    }
    tests_by_user = {
        row['user_id']: row
        for row in test_results.values('user_id').annotate(
            taken=Count('id'),
            passed=Count('id', filter=Q(passed=True)),
            avg_score=Avg('score'),
        )
    }
    tasks_by_user = {
        row['assignment__user_id']: row
        for row in submissions.values('assignment__user_id').annotate(
            submitted=Count('id'),
            approved=Count('id', filter=Q(status=TaskSubmission.Status.APPROVED)),
        )
    }
    summary['active_learners'] = sum(
        1 for row in enrollment_by_user.values() if (row.get('assigned') or 0) > 0
    )

    employee_rows = []
    for user in users_qs.select_related('position', 'specialty').order_by('last_name', 'first_name'):
        enroll_row = enrollment_by_user.get(user.id, {})
        test_row = tests_by_user.get(user.id, {})
        task_row = tasks_by_user.get(user.id, {})
        taken = test_row.get('taken') or 0
        passed = test_row.get('passed') or 0
        pass_rate = round(passed / taken * 100, 1) if taken else 0
        employee_rows.append(
            {
                'user': user,
                'assigned': enroll_row.get('assigned') or 0,
                'completed': enroll_row.get('completed') or 0,
                'in_progress': enroll_row.get('in_progress') or 0,
                'overdue': enroll_row.get('overdue') or 0,
                'progress_avg': enroll_row.get('progress_avg'),
                'tests_taken': taken,
                'avg_score': test_row.get('avg_score'),
                'pass_rate': pass_rate,
                'tasks_submitted': task_row.get('submitted') or 0,
                'tasks_approved': task_row.get('approved') or 0,
            }
        )

    enrollment_by_dept = {
        (row['user__department'] or '—'): row
        for row in enrollments.values('user__department').annotate(
            assigned=Count('id'),
            completed=Count('id', filter=Q(status=Enrollment.Status.COMPLETED)),
            in_progress=Count('id', filter=Q(status=Enrollment.Status.IN_PROGRESS)),
            overdue=Count('id', filter=Q(status=Enrollment.Status.OVERDUE)),
            progress_avg=Avg('progress'),
        )
    }
    tests_by_dept = {
        (row['user__department'] or '—'): row
        for row in test_results.values('user__department').annotate(
            taken=Count('id'),
            passed=Count('id', filter=Q(passed=True)),
            avg_score=Avg('score'),
        )
    }
    tasks_by_dept = {
        (row['assignment__user__department'] or '—'): row
        for row in submissions.values('assignment__user__department').annotate(
            submitted=Count('id'),
            approved=Count('id', filter=Q(status=TaskSubmission.Status.APPROVED)),
        )
    }
    department_rows = []
    departments = set(enrollment_by_dept) | set(tests_by_dept) | set(tasks_by_dept)
    for dept in sorted(departments):
        enroll_row = enrollment_by_dept.get(dept, {})
        test_row = tests_by_dept.get(dept, {})
        task_row = tasks_by_dept.get(dept, {})
        taken = test_row.get('taken') or 0
        passed = test_row.get('passed') or 0
        pass_rate = round(passed / taken * 100, 1) if taken else 0
        department_rows.append(
            {
                'department': dept or '—',
                'assigned': enroll_row.get('assigned') or 0,
                'completed': enroll_row.get('completed') or 0,
                'in_progress': enroll_row.get('in_progress') or 0,
                'overdue': enroll_row.get('overdue') or 0,
                'progress_avg': enroll_row.get('progress_avg'),
                'tests_taken': taken,
                'avg_score': test_row.get('avg_score'),
                'pass_rate': pass_rate,
                'tasks_submitted': task_row.get('submitted') or 0,
                'tasks_approved': task_row.get('approved') or 0,
            }
        )

    monthly_stats = list(
        test_results.annotate(month=TruncMonth('completed_at'))
        .values('month')
        .annotate(count=Count('id'), avg_score=Avg('score'))
        .order_by('month')
    )
    max_month_count = max((row['count'] for row in monthly_stats), default=0)
    for row in monthly_stats:
        row['percent'] = round(row['count'] / max_month_count * 100, 1) if max_month_count else 0

    monthly_series = [
        {
            'label': row['month'].strftime('%m.%Y') if row['month'] else '',
            'count': row['count'],
            'avg_score': float(row['avg_score']) if row['avg_score'] is not None else None,
        }
        for row in monthly_stats
    ]

    dept_series = [
        {
            'label': row['department'] or '—',
            'pass_rate': row['pass_rate'],
            'avg_score': float(row['avg_score']) if row['avg_score'] is not None else None,
            'tests_taken': row['tests_taken'],
        }
        for row in department_rows
      ]
    status_series = [
          {'label': 'Завершены', 'value': summary['courses_completed']},
          {'label': 'В процессе', 'value': summary['courses_in_progress']},
          {'label': 'Просрочены', 'value': summary['courses_overdue']},
      ]
    tasks_series = [
          {'label': 'Отправлены', 'value': summary['tasks_submitted']},
          {'label': 'Приняты', 'value': summary['tasks_approved']},
      ]

    competencies_rows = list(
        UserCompetency.objects.filter(user__in=users_qs)
        .values('competency__name')
        .annotate(avg_level=Avg('level'), count=Count('id'))
        .order_by('avg_level')[:8]
    )

    recent_tests = list(
        test_results.select_related('user', 'test')
        .order_by('-completed_at')[:10]
    )
    recent_tasks = list(
        submissions.select_related('assignment__user', 'assignment__task')
        .order_by('-submitted_at')[:10]
    )

    return {
        'summary': summary,
        'employee_rows': employee_rows,
        'department_rows': department_rows,
        'monthly_stats': monthly_stats,
        'monthly_series': monthly_series,
          'dept_series': dept_series,
          'status_series': status_series,
          'tasks_series': tasks_series,
        'competencies_rows': competencies_rows,
        'recent_tests': recent_tests,
        'recent_tasks': recent_tasks,
        'users_qs': users_qs,
        'courses_qs': courses_qs,
    }


def _test_is_assigned(test):
    if TestResult.objects.filter(test=test).exists():
        return True
    if test.is_published:
        return Enrollment.objects.filter(course=test.course).exclude(
            status=Enrollment.Status.CANCELLED
        ).exists()
    return False


def _task_has_submissions(task):
    return TaskSubmission.objects.filter(assignment__task=task).exists()


def _lesson_has_test(lesson):
    try:
        pages = lesson.pages or []
    except AttributeError:
        return False
    for page in pages:
        for block in page.get('blocks', []):
            if block.get('type') == 'test' and block.get('test_id'):
                return True
    return False


def _clone_test_with_questions(source_test, overrides=None, created_by=None, return_map=False):
    overrides = overrides or {}
    new_test = Test.objects.create(
        course=overrides.get('course') or source_test.course,
        title=overrides.get('title') or source_test.title,
        description=overrides.get('description') if overrides.get('description') is not None else source_test.description,
        passing_score=overrides.get('passing_score') if overrides.get('passing_score') is not None else source_test.passing_score,
        attempts=overrides.get('attempts') if overrides.get('attempts') is not None else source_test.attempts,
        warning_threshold=overrides.get('warning_threshold') if overrides.get('warning_threshold') is not None else source_test.warning_threshold,
        success_threshold=overrides.get('success_threshold') if overrides.get('success_threshold') is not None else source_test.success_threshold,
        retake_requires_new_attempt=overrides.get('retake_requires_new_attempt') if overrides.get('retake_requires_new_attempt') is not None else source_test.retake_requires_new_attempt,
        due_date=overrides.get('due_date') if overrides.get('due_date') is not None else source_test.due_date,
        is_published=overrides.get('is_published') if overrides.get('is_published') is not None else source_test.is_published,
        created_by=created_by or source_test.created_by,
    )
    question_map = {}
    for question in source_test.questions.prefetch_related('answers', 'matching_pairs', 'ordering_items'):
        new_question = Question.objects.create(
            test=new_test,
            type=question.type,
            text=question.text,
            points=question.points,
            image=question.image,
        )
        question_map[question.id] = new_question.id
        answers = [
            Answer(question=new_question, text=ans.text, is_correct=ans.is_correct)
            for ans in question.answers.all()
        ]
        if answers:
            Answer.objects.bulk_create(answers)
        pairs = [
            MatchingPair(
                question=new_question,
                left_text=pair.left_text,
                right_text=pair.right_text,
                order=pair.order,
            )
            for pair in question.matching_pairs.all()
        ]
        if pairs:
            MatchingPair.objects.bulk_create(pairs)
        ordering_items = [
            OrderingItem(
                question=new_question,
                text=item.text,
                position=item.position,
            )
            for item in question.ordering_items.all()
        ]
        if ordering_items:
            OrderingItem.objects.bulk_create(ordering_items)
    if return_map:
        return new_test, question_map
    return new_test


@login_required
def redirect_after_login(request):
    role = request.user.role
    if request.user.is_superuser or role == 'ADMIN':
        return redirect('/admin-panel/')
    if role == 'MANAGER':
        return redirect('/manager/')
    elif role == 'ANALYST':
        return redirect('/analytics/')
    elif role == 'EMPLOYEE':
        return redirect('/employee/')
    else:
        return redirect('/admin/')


@login_required
def profile_view(request):
    saved = False
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            user = form.save(commit=False)
            remove_avatar = bool(form.cleaned_data.get('remove_avatar'))
            has_new_avatar = 'avatar' in request.FILES
            if remove_avatar and not has_new_avatar:
                try:
                    if user.avatar:
                        user.avatar.delete(save=False)
                except Exception:
                    pass
                user.avatar = None
            user.save()
            saved = True
    else:
        form = UserProfileForm(instance=request.user)

    return render(
        request,
        'users/profile.html',
        {'form': form, 'saved': saved},
    )


@login_required
def admin_panel_view(request):
    user = request.user
    if not (user.is_superuser or user.role == 'ADMIN'):
        raise PermissionDenied
    refresh = RefreshToken.for_user(user)
    access = refresh.access_token
    return render(
        request,
        'admin_panel/dashboard.html',
        {
            'access_token': str(access),
            'refresh_token': str(refresh),
            'admin_user_id': user.id,
            'admin_email': user.email,
        },
    )


@login_required
def registration_requests_view(request):
    user = request.user
    if not (getattr(user, 'is_superuser', False) or user.role == 'ADMIN'):
        raise PermissionDenied

    status_filter = (request.GET.get('status') or 'PENDING').strip().upper()
    allowed = {'PENDING', 'APPROVED', 'REJECTED', 'ALL'}
    if status_filter not in allowed:
        status_filter = 'PENDING'

    qs = User.objects.filter(is_superuser=False, role='EMPLOYEE').order_by('-date_joined', '-id')
    if status_filter != 'ALL':
        qs = qs.filter(registration_status=status_filter)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip().lower()
        user_id = (request.POST.get('user_id') or '').strip()
        comment = (request.POST.get('comment') or '').strip()

        if action in {'approve', 'reject'} and user_id.isdigit():
            target = (
                User.objects.filter(is_superuser=False, role='EMPLOYEE', id=int(user_id))
                .first()
            )
            if target:
                now = timezone.now()
                if action == 'approve':
                    target.is_active = True
                    target.registration_status = User.RegistrationStatus.APPROVED
                else:
                    target.is_active = False
                    target.registration_status = User.RegistrationStatus.REJECTED
                target.registration_reviewed_at = now
                target.registration_reviewed_by = request.user
                target.registration_review_comment = comment
                target.save(
                    update_fields=[
                        'is_active',
                        'registration_status',
                        'registration_reviewed_at',
                        'registration_reviewed_by',
                        'registration_review_comment',
                    ]
                )
        return redirect(reverse('registration_requests') + f'?status={status_filter}')

    counts = (
        User.objects.filter(is_superuser=False, role='EMPLOYEE')
        .values('registration_status')
        .annotate(count=Count('id'))
    )
    counts_map = {row['registration_status']: row['count'] for row in counts}
    return render(
        request,
        'users/registration_requests.html',
        {
            'requests': list(qs),
            'status_filter': status_filter,
            'counts': {
                'PENDING': counts_map.get('PENDING', 0),
                'APPROVED': counts_map.get('APPROVED', 0),
                'REJECTED': counts_map.get('REJECTED', 0),
            },
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_dashboard(request):

    return render(request, 'users/manager_dashboard.html')


@login_required
@role_required(['MANAGER'])
def manager_course_list(request):
    courses = (
        Course.objects.filter(created_by=request.user)
        .select_related('specialty')
        .annotate(
            enrollments_count=Count('enrollments', distinct=True),
            tests_count=Count('tests', distinct=True),
            tasks_count=Count('tasks', distinct=True),
        )
        .order_by('-created_at')
    )
    return render(request, 'manager/course_list.html', {'courses': courses})


@login_required
@role_required(['MANAGER'])
def manager_course_detail(request, course_id):
    course = get_object_or_404(
        Course.objects.select_related('specialty', 'created_by'),
        id=course_id,
        created_by=request.user,
    )
    enrollments = (
        Enrollment.objects.filter(course=course)
        .select_related('user', 'user__specialty', 'user__position')
        .order_by('user__last_name', 'user__first_name')
    )
    tests = (
        Test.objects.filter(course=course)
        .annotate(question_count=Count('questions'))
        .order_by('-created_at')
    )
    tasks = Task.objects.filter(course=course).order_by('-created_at')
    lessons = list(Lesson.objects.filter(course=course).order_by('order', 'id'))
    for item in lessons:
        item.has_test = _lesson_has_test(item)
    tests_with_results = set(
        TestResult.objects.filter(test__in=tests)
        .values_list('test_id', flat=True)
        .distinct()
    )
    has_active_enrollments = Enrollment.objects.filter(course=course).exclude(
        status=Enrollment.Status.CANCELLED
    ).exists()
    test_results_summary = {}
    if tests:
        results_rows = (
            TestResult.objects.filter(test__in=tests)
            .values('test_id', 'status')
            .annotate(count=Count('id'))
        )
        for test in tests:
            test_results_summary[test.id] = {
                'PASSED': 0,
                'FAILED': 0,
                'UNDER_REVIEW': 0,
                'RETURNED': 0,
                'DECLINED': 0,
                'TOTAL': 0,
            }
        for row in results_rows:
            summary = test_results_summary.get(row['test_id'])
            if not summary:
                continue
            status_value = row['status']
            summary[status_value] = row['count']
            summary['TOTAL'] += row['count']
    tests_data = []
    for test in tests:
        assigned = test.id in tests_with_results or (test.is_published and has_active_enrollments)
        tests_data.append(
            {
                'test': test,
                'summary': test_results_summary.get(test.id),
                'assigned': assigned,
            }
        )
    return render(
        request,
        'manager/course_detail.html',
        {
            'course': course,
            'enrollments': enrollments,
            'tests': tests,
            'tests_data': tests_data,
            'tasks': tasks,
            'lessons': lessons,
        },
    )


@login_required
@role_required(['EMPLOYEE'])
def employee_dashboard(request):
    enrollments = (
        Enrollment.objects.filter(user=request.user)
        .select_related('course', 'course__specialty', 'course__created_by')
        .order_by('-assigned_at')
    )
    today = timezone.localdate()
    enrollments_view = []
    overdue_count = 0
    due_soon_count = 0
    for enrollment in enrollments:
        due_date = enrollment.due_date
        days_left = None
        due_status = 'none'
        due_text = '—'
        if due_date:
            days_left = (due_date - today).days
            if days_left < 0:
                due_status = 'overdue'
                overdue_count += 1
                due_text = f"Просрочено на {abs(days_left)} дн. ({due_date:%d.%m.%Y})"
            elif days_left <= 3:
                due_status = 'soon'
                due_soon_count += 1
                due_text = f"Срок {due_date:%d.%m.%Y} (через {days_left} дн.)"
            else:
                due_status = 'ok'
                due_text = due_date.strftime('%d.%m.%Y')
        enrollments_view.append(
            {
                'enrollment': enrollment,
                'due_status': due_status,
                'due_text': due_text,
                'days_left': days_left,
            }
        )

    return render(
        request,
        'employee/dashboard.html',
        {
            'enrollments': enrollments,
            'enrollments_view': enrollments_view,
            'overdue_count': overdue_count,
            'due_soon_count': due_soon_count,
        },
    )


def _build_user_progress_context(target_user):
    today = timezone.localdate()
    enrollments = (
        Enrollment.objects.filter(user=target_user)
        .select_related('course', 'course__specialty')
        .order_by('-assigned_at', '-id')
    )
    courses_progress = []
    overdue_items = []
    due_soon_items = []
    for enrollment in enrollments:
        courses_progress.append(
            {
                'course_id': enrollment.course_id,
                'course_title': enrollment.course.title,
                'progress': float(enrollment.progress or 0),
                'status': enrollment.status,
                'completed': bool(enrollment.completed),
            }
        )
        if enrollment.due_date:
            days_left = (enrollment.due_date - today).days
            item = {
                'title': enrollment.course.title,
                'due_date': enrollment.due_date,
                'days_left': days_left,
                'days_overdue': abs(days_left) if days_left < 0 else 0,
            }
            if days_left < 0 and not enrollment.completed:
                overdue_items.append(item)
            elif 0 <= days_left <= 3 and not enrollment.completed:
                due_soon_items.append(item)

    test_monthly_rows = list(
        TestResult.objects.filter(user=target_user)
        .exclude(status=TestResult.Status.UNDER_REVIEW)
        .annotate(month=TruncMonth('completed_at'))
        .values('month')
        .annotate(count=Count('id'), avg_score=Avg('score'))
        .order_by('month')
    )
    test_monthly_series = [
        {
            'month': row['month'].strftime('%Y-%m'),
            'count': int(row['count'] or 0),
            'avg_score': float(row['avg_score']) if row['avg_score'] is not None else None,
        }
        for row in test_monthly_rows
        if row.get('month') is not None
    ]

    task_monthly_rows = list(
        TaskSubmission.objects.filter(assignment__user=target_user)
        .annotate(month=TruncMonth('submitted_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    task_monthly_series = [
        {
            'month': row['month'].strftime('%Y-%m'),
            'count': int(row['count'] or 0),
        }
        for row in task_monthly_rows
        if row.get('month') is not None
    ]

    competency_rows = [
        {
            'name': item.competency.name,
            'level': int(item.level),
            'source': item.source,
            'updated_at': item.updated_at,
        }
        for item in UserCompetency.objects.filter(user=target_user).select_related('competency').order_by(
            '-updated_at',
            'competency__name',
        )
    ]

    enrollments_summary = enrollments.aggregate(
        assigned=Count('id'),
        completed=Count('id', filter=Q(completed=True)),
        overdue=Count('id', filter=Q(status=Enrollment.Status.OVERDUE)),
        in_progress=Count('id', filter=Q(status=Enrollment.Status.IN_PROGRESS)),
        progress_avg=Avg('progress'),
    )
    tests_summary = (
        TestResult.objects.filter(user=target_user)
        .exclude(score__isnull=True)
        .aggregate(
            avg_score=Avg('score'),
            tests_taken=Count('id'),
            tests_passed=Count('id', filter=Q(status=TestResult.Status.PASSED)),
        )
    )

    return {
        'target_user': target_user,
        'summary': {
            'assigned_courses': enrollments_summary.get('assigned') or 0,
            'completed_courses': enrollments_summary.get('completed') or 0,
            'courses_in_progress': enrollments_summary.get('in_progress') or 0,
            'courses_overdue': enrollments_summary.get('overdue') or 0,
            'progress_avg': (
                float(enrollments_summary['progress_avg'])
                if enrollments_summary.get('progress_avg') is not None
                else None
            ),
            'tests_taken': tests_summary.get('tests_taken') or 0,
            'tests_passed': tests_summary.get('tests_passed') or 0,
            'avg_score': float(tests_summary['avg_score']) if tests_summary.get('avg_score') is not None else None,
        },
        'courses_progress': courses_progress,
        'test_monthly_series': test_monthly_series,
        'task_monthly_series': task_monthly_series,
        'competency_rows': competency_rows,
        'overdue_items': sorted(overdue_items, key=lambda x: x['days_left']),
        'due_soon_items': sorted(due_soon_items, key=lambda x: x['days_left']),
    }


@login_required
@role_required(['EMPLOYEE'])
def employee_progress(request):
    context = _build_user_progress_context(request.user)
    return render(request, 'users/progress.html', context)


@login_required
@role_required(['MANAGER', 'ANALYST', 'ADMIN'])
def manager_student_progress(request, student_id):
    student = get_object_or_404(User, id=student_id, role='EMPLOYEE')
    context = _build_user_progress_context(student)
    context['view_as'] = 'manager'
    return render(request, 'users/progress.html', context)


@login_required
@role_required(['EMPLOYEE'])
def employee_course_detail(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    enrollment = get_object_or_404(Enrollment, course=course, user=request.user)

    tests = (
        Test.objects.filter(course=course, is_published=True)
        .annotate(question_count=Count('questions'))
        .order_by('-created_at')
    )
    results = (
        TestResult.objects.filter(user=request.user, test__in=tests)
        .order_by('test_id', '-completed_at', '-id')
    )
    latest_results = {}
    for result in results:
        latest_results.setdefault(result.test_id, result)

    attempt_counts = (
        TestResult.objects.filter(user=request.user, test__in=tests)
        .values('test_id')
        .annotate(count=Count('id'))
    )
    attempts_by_test = {}
    for row in attempt_counts:
        test_id = row['test_id']
        attempts_by_test[test_id] = row['count']

    assignments = (
        TaskAssignment.objects.filter(
            user=request.user,
            task__course=course,
            task__is_published=True,
        )
        .select_related('task')
        .order_by('-assigned_at')
    )
    lessons = list(Lesson.objects.filter(course=course, is_published=True).order_by('order', 'id'))
    test_to_lesson = {}
    for item in lessons:
        item.has_test = _lesson_has_test(item)
        pages = item.pages or []
        for page in pages:
            for block in page.get('blocks', []):
                if block.get('type') == 'test' and block.get('test_id') and block.get('test_id') not in test_to_lesson:
                    test_to_lesson[block.get('test_id')] = item

    tests_available = []
    tests_completed = []
    tests_retake = []
    for test in tests:
        attempts_used = attempts_by_test.get(test.id, 0)
        display_result = latest_results.get(test.id)
        retake_accepted = (
            display_result
            and display_result.status == TestResult.Status.RETURNED
            and display_result.retake_accepted
        )
        extra_attempts = 1 if retake_accepted and test.retake_requires_new_attempt else 0
        attempts_total = test.attempts + extra_attempts
        can_attempt = attempts_used < test.attempts
        if display_result and display_result.status == TestResult.Status.RETURNED:
            if not retake_accepted:
                can_attempt = False
            elif test.retake_requires_new_attempt:
                can_attempt = attempts_used < attempts_total
            else:
                can_attempt = True
        if display_result and display_result.status == TestResult.Status.DECLINED:
            can_attempt = False
        label = 'Пройти ещё раз' if attempts_used > 0 else 'Пройти'
        item = {
            'test': test,
            'result': display_result,
            'attempts_used': attempts_used,
            'attempts_total': attempts_total,
            'can_attempt': can_attempt,
            'button_label': label,
            'retake_accepted': retake_accepted,
            'needs_retake_decision': (
                display_result and display_result.status == TestResult.Status.RETURNED and not retake_accepted
            ),
            'lesson': test_to_lesson.get(test.id),
        }
        if display_result and display_result.status == TestResult.Status.RETURNED:
            tests_retake.append(item)
        elif display_result:
            tests_completed.append(item)
        else:
            tests_available.append(item)

    return render(
        request,
        'employee/course_detail.html',
        {
            'course': course,
            'enrollment': enrollment,
            'tests_available': tests_available,
            'tests_completed': tests_completed,
            'tests_retake': tests_retake,
            'assignments': assignments,
            'lessons': lessons,
        },
    )


@login_required
@role_required(['EMPLOYEE'])
def employee_lesson_detail(request, lesson_id):
    lesson = get_object_or_404(
        Lesson.objects.select_related('course'),
        id=lesson_id,
    )
    if not Enrollment.objects.filter(course=lesson.course, user=request.user).exists():
        raise PermissionDenied
    pages = lesson.pages or []
    return render(
        request,
        'employee/lesson_detail.html',
        {
            'lesson': lesson,
            'pages': pages,
        },
    )


@login_required
@role_required(['EMPLOYEE'])
def employee_test_decline(request, test_id):
    if request.method != 'POST':
        raise PermissionDenied

    test = get_object_or_404(Test.objects.select_related('course'), id=test_id, is_published=True)
    if not Enrollment.objects.filter(course=test.course, user=request.user).exists():
        raise PermissionDenied

    result = (
        TestResult.objects.filter(
            user=request.user,
            test=test,
            status=TestResult.Status.RETURNED,
        )
        .order_by('-completed_at', '-id')
        .first()
    )
    if result:
        result.status = TestResult.Status.DECLINED
        result.score = None
        result.passed = None
        result.save(update_fields=['status', 'score', 'passed'])

    return redirect('employee_course_detail', course_id=test.course_id)


@login_required
@role_required(['EMPLOYEE'])
def employee_test_accept(request, test_id):
    if request.method != 'POST':
        raise PermissionDenied

    test = get_object_or_404(Test.objects.select_related('course'), id=test_id, is_published=True)
    if not Enrollment.objects.filter(course=test.course, user=request.user).exists():
        raise PermissionDenied

    result = (
        TestResult.objects.filter(
            user=request.user,
            test=test,
            status=TestResult.Status.RETURNED,
        )
        .order_by('-completed_at', '-id')
        .first()
    )
    accepted = False
    if result and not result.retake_accepted:
        result.retake_accepted = True
        result.save(update_fields=['retake_accepted'])
        accepted = True

    if accepted:
        return redirect('employee_test_take', test_id=test.id)
    return redirect('employee_course_detail', course_id=test.course_id)


@login_required
@role_required(['EMPLOYEE'])
def employee_test_take(request, test_id):
    test = get_object_or_404(Test.objects.select_related('course'), id=test_id, is_published=True)
    if not Enrollment.objects.filter(course=test.course, user=request.user).exists():
        raise PermissionDenied

    questions = list(
        test.questions.all()
        .prefetch_related('answers', 'matching_pairs', 'ordering_items')
        .order_by('id')
    )
    questions_data = []
    for question in questions:
        item = {'question': question, 'type': question.type}
        if question.type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
            answers = list(question.answers.all())
            correct_ids = {answer.id for answer in answers if answer.is_correct}
            item.update(
                {
                    'answers': answers,
                    'correct_ids': correct_ids,
                    'multi': question.type == QUESTION_TYPE_MULTI,
                }
            )
        elif question.type == QUESTION_TYPE_MATCHING:
            pairs = list(question.matching_pairs.all().order_by('order', 'id'))
            right_options = [{'id': pair.id, 'text': pair.right_text} for pair in pairs]
            item.update({'pairs': pairs, 'right_options': right_options})
        elif question.type == QUESTION_TYPE_ORDERING:
            ordering_items = list(question.ordering_items.all().order_by('position', 'id'))
            item.update({'ordering_items': ordering_items})
        questions_data.append(item)
    total_points = sum(item['question'].points for item in questions_data)
    manual_exists = any(question.type in MANUAL_QUESTION_TYPES for question in questions)
    expected_evaluation = (
        Test.EvaluationType.MANUAL if manual_exists else Test.EvaluationType.AUTO
    )
    if test.evaluation_type != expected_evaluation:
        Test.objects.filter(pk=test.pk).update(evaluation_type=expected_evaluation)
        test.evaluation_type = expected_evaluation

    attempts_used = TestResult.objects.filter(
        user=request.user,
        test=test,
    ).count()
    returned_result = (
        TestResult.objects.filter(
            user=request.user,
            test=test,
            status=TestResult.Status.RETURNED,
        )
        .order_by('-completed_at', '-id')
        .first()
    )
    declined_result = (
        TestResult.objects.filter(
            user=request.user,
            test=test,
            status=TestResult.Status.DECLINED,
        )
        .order_by('-completed_at', '-id')
        .first()
    )
    extra_attempts = (
        1
        if returned_result and returned_result.retake_accepted and test.retake_requires_new_attempt
        else 0
    )
    attempts_total = test.attempts + extra_attempts
    attempts_left = max(attempts_total - attempts_used, 0)
    retake_pending = bool(returned_result and not returned_result.retake_accepted)

    if request.method == 'POST':
        if declined_result:
            return render(
                request,
                'employee/test_take.html',
                {
                    'test': test,
                    'questions_data': questions_data,
                    'error': 'Вы отказались от пересдачи. Повторная попытка недоступна.',
                    'attempts_used': attempts_used,
                    'attempts_left': attempts_left,
                    'attempts_total': attempts_total,
                    'block_attempt': True,
                },
            )
        if retake_pending:
            return render(
                request,
                'employee/test_take.html',
                {
                    'test': test,
                    'questions_data': questions_data,
                    'error': 'Сначала примите пересдачу.',
                    'attempts_used': attempts_used,
                    'attempts_left': attempts_left,
                    'attempts_total': attempts_total,
                    'block_attempt': True,
                },
            )
        correct_points = 0
        missing = []
        selected_rows = []
        extra_answers = []
        for item in questions_data:
            question = item['question']
            question_type = question.type
            key = f'question_{question.id}'
            if question_type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
                selected_ids = request.POST.getlist(key)
                if not selected_ids:
                    missing.append(question.id)
                    continue
                try:
                    selected_ids = {int(value) for value in selected_ids}
                except ValueError:
                    return render(
                        request,
                        'employee/test_take.html',
                        {
                            'test': test,
                            'questions_data': questions_data,
                            'error': 'Некорректный ответ.',
                            'attempts_used': attempts_used,
                            'attempts_left': attempts_left,
                        },
                    )
                if question_type == QUESTION_TYPE_SINGLE and len(selected_ids) != 1:
                    return render(
                        request,
                        'employee/test_take.html',
                        {
                            'test': test,
                            'questions_data': questions_data,
                            'error': 'Некорректный ответ.',
                            'attempts_used': attempts_used,
                            'attempts_left': attempts_left,
                        },
                    )
                answer_map = {answer.id: answer.is_correct for answer in item.get('answers', [])}
                for answer_id in selected_ids:
                    if answer_id not in answer_map:
                        return render(
                            request,
                            'employee/test_take.html',
                            {
                                'test': test,
                                'questions_data': questions_data,
                                'error': 'Некорректный ответ.',
                                'attempts_used': attempts_used,
                                'attempts_left': attempts_left,
                            },
                        )
                    selected_rows.append(
                        {
                            'question_id': question.id,
                            'answer_id': answer_id,
                            'is_correct': answer_map[answer_id],
                        }
                    )
                correct_ids = item.get('correct_ids', set())
                if correct_ids:
                    if question_type == QUESTION_TYPE_MULTI:
                        if selected_ids == correct_ids:
                            correct_points += question.points
                    else:
                        if next(iter(selected_ids)) in correct_ids:
                            correct_points += question.points
            elif question_type == QUESTION_TYPE_MATCHING:
                pairs = item.get('pairs', [])
                if not pairs:
                    missing.append(question.id)
                    continue
                pair_ids = {pair.id for pair in pairs}
                mapping = []
                any_selected = False
                missing_pair = False
                for pair in pairs:
                    field_name = f'match_{question.id}_{pair.id}'
                    selected = request.POST.get(field_name)
                    if selected is None or selected == '':
                        missing_pair = True
                        continue
                    any_selected = True
                    try:
                        selected_id = int(selected)
                    except ValueError:
                        return render(
                            request,
                            'employee/test_take.html',
                            {
                                'test': test,
                                'questions_data': questions_data,
                                'error': 'Некорректный ответ.',
                                'attempts_used': attempts_used,
                                'attempts_left': attempts_left,
                            },
                        )
                    if selected_id not in pair_ids:
                        return render(
                            request,
                            'employee/test_take.html',
                            {
                                'test': test,
                                'questions_data': questions_data,
                                'error': 'Некорректный ответ.',
                                'attempts_used': attempts_used,
                                'attempts_left': attempts_left,
                            },
                        )
                    mapping.append({'left_id': pair.id, 'right_id': selected_id})
                if missing_pair:
                    missing.append(question.id)
                is_correct = (
                    len(mapping) == len(pair_ids)
                    and all(item['left_id'] == item['right_id'] for item in mapping)
                )
                if is_correct:
                    correct_points += question.points
                if any_selected:
                    extra_answers.append(
                        TestAnswer(
                            question_id=question.id,
                            answer_data={'matches': mapping},
                            is_correct=is_correct if mapping else None,
                        )
                    )
            elif question_type == QUESTION_TYPE_ORDERING:
                ordering_items = item.get('ordering_items', [])
                if not ordering_items:
                    missing.append(question.id)
                    continue
                positions = {}
                missing_position = False
                any_selected = False
                for ordering_item in ordering_items:
                    field_name = f'order_{question.id}_{ordering_item.id}'
                    value = request.POST.get(field_name)
                    if value is None or value.strip() == '':
                        missing_position = True
                        continue
                    any_selected = True
                    try:
                        pos = int(value)
                    except ValueError:
                        return render(
                            request,
                            'employee/test_take.html',
                            {
                                'test': test,
                                'questions_data': questions_data,
                                'error': 'Некорректный ответ.',
                                'attempts_used': attempts_used,
                                'attempts_left': attempts_left,
                            },
                        )
                    positions[ordering_item.id] = pos
                if missing_position:
                    missing.append(question.id)
                order_list = []
                is_correct = False
                if positions:
                    order_list = [
                        item_id
                        for item_id, _ in sorted(positions.items(), key=lambda row: (row[1], row[0]))
                    ]
                    if len(positions) == len(ordering_items):
                        pos_values = list(positions.values())
                        expected_positions = list(range(1, len(ordering_items) + 1))
                        if len(set(pos_values)) == len(ordering_items) and sorted(pos_values) == expected_positions:
                            correct_order = [item.id for item in ordering_items]
                            if order_list == correct_order:
                                is_correct = True
                if is_correct:
                    correct_points += question.points
                if any_selected:
                    extra_answers.append(
                        TestAnswer(
                            question_id=question.id,
                            answer_data={
                                'order': order_list,
                                'positions': [
                                    {'id': item_id, 'position': position}
                                    for item_id, position in positions.items()
                                ],
                            },
                            is_correct=is_correct if positions else None,
                        )
                    )
            elif question_type in {QUESTION_TYPE_SHORT, QUESTION_TYPE_LONG}:
                text = request.POST.get(key, '').strip()
                attachment = None
                if question_type == QUESTION_TYPE_LONG:
                    attachment = request.FILES.get(f'attachment_{question.id}')
                if text == '':
                    missing.append(question.id)
                else:
                    extra_answers.append(
                        TestAnswer(
                            question_id=question.id,
                            answer_text=text,
                            attachment=attachment,
                            is_correct=None,
                        )
                    )
            else:
                return render(
                    request,
                    'employee/test_take.html',
                    {
                        'test': test,
                        'questions_data': questions_data,
                        'error': 'Некорректный ответ.',
                        'attempts_used': attempts_used,
                        'attempts_left': attempts_left,
                        'attempts_total': attempts_total,
                    },
                )

        reuse_returned = returned_result and not test.retake_requires_new_attempt
        if reuse_returned:
            attempt_number = returned_result.attempt_number
        else:
            attempt_limit = test.attempts
            if returned_result and returned_result.retake_accepted and test.retake_requires_new_attempt:
                attempt_limit += 1
            if attempts_used >= attempt_limit:
                return render(
                    request,
                    'employee/test_take.html',
                    {
                        'test': test,
                        'questions_data': questions_data,
                        'error': 'Лимит попыток исчерпан.',
                        'attempts_used': attempts_used,
                        'attempts_left': attempts_left,
                    },
                )
            attempt_number = attempts_used + 1

        if manual_exists:
            status_value = TestResult.Status.UNDER_REVIEW
            score = None
            passed = None
        else:
            score = int(round((correct_points / total_points) * 100)) if total_points > 0 else 0
            passed = score >= test.passing_score
            if missing:
                passed = False
            status_value = TestResult.Status.PASSED if passed else TestResult.Status.FAILED

        if reuse_returned:
            result = returned_result
            result.score = score
            result.passed = passed
            result.status = status_value
            result.attempt_number = attempt_number
            result.save(update_fields=['score', 'passed', 'status', 'attempt_number'])
            TestAnswer.objects.filter(test_result=result).delete()
            Feedback.objects.filter(test_result=result).delete()
        else:
            result = TestResult.objects.create(
                user=request.user,
                test=test,
                score=score,
                passed=passed,
                status=status_value,
                attempt_number=attempt_number,
            )
        test_answers = []
        if selected_rows:
            test_answers.extend(
                [
                    TestAnswer(
                        test_result=result,
                        question_id=row['question_id'],
                        answer_id=row['answer_id'],
                        is_correct=row['is_correct'],
                    )
                    for row in selected_rows
                ]
            )
        if extra_answers:
            for answer in extra_answers:
                answer.test_result = result
            test_answers.extend(extra_answers)
        if test_answers:
            answers_with_files = [answer for answer in test_answers if answer.attachment]
            answers_without_files = [answer for answer in test_answers if not answer.attachment]
            if answers_without_files:
                TestAnswer.objects.bulk_create(answers_without_files)
            for answer in answers_with_files:
                answer.save()
        TrainingEvent.objects.create(
            user=request.user,
            event_type='TEST_COMPLETED',
            course=test.course,
            test=test,
        )
        return redirect('employee_test_result', result_id=result.id)

    return render(
        request,
        'employee/test_take.html',
        {
            'test': test,
            'questions_data': questions_data,
            'attempts_used': attempts_used,
            'attempts_left': attempts_left,
            'attempts_total': attempts_total,
            'block_attempt': bool(declined_result or retake_pending),
            'error': (
                'Вы отказались от пересдачи. Повторная попытка недоступна.'
                if declined_result
                else 'Сначала примите пересдачу.' if retake_pending else None
            ),
        },
    )


@login_required
@role_required(['EMPLOYEE'])
def employee_test_result(request, result_id):
    test_result = get_object_or_404(
        TestResult.objects.select_related('test', 'test__course'),
        id=result_id,
        user=request.user,
    )
    test_feedback = Feedback.objects.filter(test_result=test_result).order_by('-created_at').first()
    questions = (
        Question.objects.filter(test=test_result.test)
        .prefetch_related('answers', 'matching_pairs', 'ordering_items')
        .order_by('id')
    )
    selections = (
        TestAnswer.objects.filter(test_result=test_result)
        .select_related('answer', 'question')
    )
    answers_by_question = {}
    for sel in selections:
        answers_by_question.setdefault(sel.question_id, []).append(sel)

    result_questions = []
    for question in questions:
        question_type = question.type
        item = {'question': question, 'type': question_type}
        if question_type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
            answers = list(question.answers.all())
            correct_ids = {answer.id for answer in answers if answer.is_correct}
            selected_ids = {
                sel.answer_id
                for sel in answers_by_question.get(question.id, [])
                if sel.answer_id
            }
            item.update(
                {
                    'answers': answers,
                    'correct_ids': correct_ids,
                    'selected_ids': selected_ids,
                }
            )
        elif question_type == QUESTION_TYPE_MATCHING:
            pairs = list(question.matching_pairs.all().order_by('order', 'id'))
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_data),
                None,
            )
            match_map = {}
            if selection and selection.answer_data:
                for match in selection.answer_data.get('matches', []):
                    left_id = match.get('left_id')
                    right_id = match.get('right_id')
                    if left_id is not None and right_id is not None:
                        match_map[left_id] = right_id
            right_by_id = {pair.id: pair.right_text for pair in pairs}
            matching_rows = []
            for pair in pairs:
                selected_id = match_map.get(pair.id)
                matching_rows.append(
                    {
                        'left_text': pair.left_text,
                        'right_text': pair.right_text,
                        'selected_right_text': right_by_id.get(selected_id),
                        'selected_right_id': selected_id,
                        'right_id': pair.id,
                    }
                )
            item.update({'matching_rows': matching_rows})
        elif question_type == QUESTION_TYPE_ORDERING:
            ordering_items = list(question.ordering_items.all().order_by('position', 'id'))
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_data),
                None,
            )
            order_list = []
            positions_map = {}
            if selection and selection.answer_data:
                order_list = selection.answer_data.get('order') or []
                for row in selection.answer_data.get('positions') or []:
                    item_id = row.get('id')
                    position = row.get('position')
                    if item_id is not None:
                        positions_map[item_id] = position
            if not positions_map and order_list:
                for idx, item_id in enumerate(order_list, start=1):
                    positions_map[item_id] = idx
            ordering_rows = [
                {
                    'id': item.id,
                    'text': item.text,
                    'position': item.position,
                    'user_position': positions_map.get(item.id),
                }
                for item in ordering_items
            ]
            item.update({'ordering_rows': ordering_rows})
        elif question_type in {QUESTION_TYPE_SHORT, QUESTION_TYPE_LONG}:
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_text is not None),
                None,
            )
            item.update(
                {
                    'answer_text': selection.answer_text if selection else '',
                    'answer_file': selection.attachment if selection else None,
                }
            )
        result_questions.append(item)

    return render(
        request,
        'employee/test_result.html',
        {
            'test_result': test_result,
            'test_feedback': test_feedback,
            'result_questions': result_questions,
            'show_review': test_result.status in {TestResult.Status.PASSED, TestResult.Status.FAILED},
        },
    )


@login_required
@role_required(['EMPLOYEE'])
def employee_task_detail(request, task_id):
    task = get_object_or_404(Task.objects.select_related('course'), id=task_id, is_published=True)
    assignment = get_object_or_404(TaskAssignment, task=task, user=request.user)
    submission = TaskSubmission.objects.filter(assignment=assignment).order_by('-submitted_at').first()
    review = None
    if submission:
        review = TaskReview.objects.filter(task_submission=submission).order_by('-created_at').first()

    if request.method == 'POST':
        form = TaskSubmissionForm(request.POST, request.FILES, instance=submission)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.assignment = assignment
            submission.status = 'SUBMITTED'
            submission.save()
            TrainingEvent.objects.create(
                user=request.user,
                event_type='TASK_SUBMITTED',
                course=task.course,
                task=task,
            )
            return redirect('employee_task_detail', task_id=task.id)
    else:
        form = TaskSubmissionForm(instance=submission)

    return render(
        request,
        'employee/task_detail.html',
        {
            'task': task,
            'assignment': assignment,
            'submission': submission,
            'form': form,
            'review': review,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_task_list(request):
    tasks = (
        Task.objects.filter(created_by=request.user)
        .select_related('course')
        .prefetch_related('assignments__user')
        .annotate(
            assignments_count=Count('assignments', distinct=True),
            submissions_count=Count('assignments__submissions', distinct=True),
        )
        .order_by('-created_at')
    )
    return render(request, 'manager/task_list.html', {'tasks': tasks})


@login_required
@role_required(['MANAGER'])
def manager_lesson_list(request):
    courses = Course.objects.filter(created_by=request.user).order_by('title')
    lessons = (
        Lesson.objects.filter(course__in=courses)
        .select_related('course')
        .order_by('course_id', 'order', 'id')
    )
    selected_course = request.GET.get('course')
    if selected_course:
        lessons = lessons.filter(course_id=selected_course)
    lessons_list = list(lessons)
    for item in lessons_list:
        item.has_test = _lesson_has_test(item)
    return render(
        request,
        'manager/lesson_list.html',
        {
            'lessons': lessons_list,
            'courses': courses,
            'selected_course': selected_course,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_lesson_builder(request, lesson_id=None):
    courses_qs = Course.objects.filter(created_by=request.user).order_by('title')
    lesson = None
    if lesson_id:
        lesson = get_object_or_404(
            Lesson.objects.select_related('course'),
            id=lesson_id,
            course__created_by=request.user,
        )
    course = lesson.course if lesson else None
    course_id = request.GET.get('course') or (course.id if course else None)
    if course_id and not course:
        course = courses_qs.filter(id=course_id).first()
    tests_qs = Test.objects.filter(course=course) if course else Test.objects.filter(course__in=courses_qs)

    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        description = (request.POST.get('description') or '').strip()
        pages_raw = request.POST.get('pages_json', '').strip() or '[]'
        order_raw = request.POST.get('order')
        is_published = request.POST.get('is_published') == 'on'
        selected_course_id = (request.POST.get('course') or '').strip() or None
        course = courses_qs.filter(id=selected_course_id).first() if selected_course_id else None
        if not course:
            return render(
                request,
                'manager/lesson_builder.html',
                {
                    'lesson': lesson,
                    'courses': courses_qs,
                    'tests': tests_qs,
                    'error': 'Выберите курс.',
                },
            )
        if not title:
            return render(
                request,
                'manager/lesson_builder.html',
                {
                    'lesson': lesson,
                    'courses': courses_qs,
                    'tests': tests_qs,
                    'error': 'Укажите название урока.',
                },
            )
        if order_raw is None:
            order = lesson.order if lesson else 0
        else:
            try:
                order = int(order_raw)
            except ValueError:
                order = 0
        try:
            import json
            pages = json.loads(pages_raw) if pages_raw else []
        except ValueError:
            pages = []

        if lesson is None:
            lesson = Lesson.objects.create(
                course=course,
                title=title,
                description=description,
                pages=pages,
                order=order,
                is_published=is_published,
                created_by=request.user,
            )
        else:
            lesson.course = course
            lesson.title = title
            lesson.description = description
            lesson.pages = pages
            lesson.order = order
            lesson.is_published = is_published
            lesson.save()
        return redirect('manager_course_detail', course_id=lesson.course_id)

    initial_pages = lesson.pages if lesson else []
    return render(
        request,
        'manager/lesson_builder.html',
        {
            'lesson': lesson,
            'courses': courses_qs,
            'tests': tests_qs,
            'initial_pages': initial_pages,
            'selected_course': course,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_lesson_delete(request, lesson_id):
    if request.method != 'POST':
        raise PermissionDenied
    lesson = get_object_or_404(
        Lesson,
        id=lesson_id,
        course__created_by=request.user,
    )
    course_id = lesson.course_id
    lesson.delete()
    return redirect('manager_course_detail', course_id=course_id)


@login_required
@role_required(['MANAGER'])
def manager_lesson_asset_upload(request):
    if request.method != 'POST':
        raise PermissionDenied
    asset_type = request.POST.get('asset_type')
    upload = request.FILES.get('file')
    if not upload or asset_type not in {LessonAsset.AssetType.IMAGE, LessonAsset.AssetType.FILE, LessonAsset.AssetType.VIDEO}:
        return JsonResponse({'error': 'Некорректный файл.'}, status=400)
    asset = LessonAsset.objects.create(
        file=upload,
        asset_type=asset_type,
        created_by=request.user,
    )
    return JsonResponse(
        {
            'id': asset.id,
            'url': asset.file.url,
            'name': asset.file.name.rsplit('/', 1)[-1],
            'type': asset.asset_type,
        }
    )


@login_required
@role_required(['MANAGER'])
def manager_test_quick_create(request):
    if request.method != 'POST':
        raise PermissionDenied
    try:
        import json
        payload = json.loads(request.body or '{}')
    except ValueError:
        payload = {}
    title = (payload.get('title') or '').strip()
    course_id = payload.get('course')
    source_test_id = payload.get('source_test_id')
    if not course_id and not source_test_id:
        return JsonResponse({'error': 'Укажите курс.'}, status=400)
    course = None
    if course_id:
        course = Course.objects.filter(id=course_id, created_by=request.user).first()
        if not course:
            return JsonResponse({'error': 'Курс не найден.'}, status=404)
    source_test = None
    if source_test_id:
        source_test = (
            Test.objects.select_related('course')
            .filter(id=source_test_id, course__created_by=request.user)
            .first()
        )
        if not source_test:
            return JsonResponse({'error': 'Тест не найден.'}, status=404)
        if course is None:
            course = source_test.course
    def _to_int(value, default):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    due_date = parse_date(payload.get('due_date')) if payload.get('due_date') else None
    if not title and source_test:
        title = source_test.title
    if not title:
        return JsonResponse({'error': 'Укажите название теста.'}, status=400)

    overrides = {
        'course': course,
        'title': title,
        'description': payload.get('description'),
        'passing_score': _to_int(payload.get('passing_score'), 70),
        'attempts': _to_int(payload.get('attempts'), 1),
        'warning_threshold': _to_int(payload.get('warning_threshold'), 50),
        'success_threshold': _to_int(payload.get('success_threshold'), 70),
        'retake_requires_new_attempt': bool(payload.get('retake_requires_new_attempt', True)),
        'due_date': due_date,
        'is_published': bool(payload.get('is_published', False)),
    }

    if source_test:
        if overrides.get('title') == source_test.title:
            base = source_test.title
            idx = 2
            while Test.objects.filter(course=source_test.course, title=f"{base} (версия {idx})").exists():
                idx += 1
            overrides['title'] = f"{base} (версия {idx})"
        test = _clone_test_with_questions(source_test, overrides=overrides, created_by=request.user)
    else:
        test = Test.objects.create(
            course=course,
            title=title,
            description=payload.get('description') or '',
            passing_score=overrides['passing_score'],
            attempts=overrides['attempts'],
            warning_threshold=overrides['warning_threshold'],
            success_threshold=overrides['success_threshold'],
            retake_requires_new_attempt=overrides['retake_requires_new_attempt'],
            due_date=overrides['due_date'],
            is_published=overrides['is_published'],
            created_by=request.user,
        )
    return JsonResponse(
        {
            'id': test.id,
            'title': test.title,
        }
    )


@login_required
@role_required(['MANAGER'])
def manager_test_quick_question(request):
    if request.method != 'POST':
        raise PermissionDenied
    payload = {}
    if request.FILES or request.POST:
        payload = request.POST.dict()
        def _load_json(key, default):
            raw = payload.get(key)
            if raw is None or raw == '':
                return default
            try:
                import json
                return json.loads(raw)
            except ValueError:
                return default
        answers = _load_json('answers', [])
        pairs = _load_json('pairs', [])
        ordering = _load_json('ordering', [])
        test_settings = _load_json('test_settings', {})
    else:
        try:
            import json
            payload = json.loads(request.body or '{}')
        except ValueError:
            payload = {}
        answers = payload.get('answers') or []
        pairs = payload.get('pairs') or []
        ordering = payload.get('ordering') or []
        test_settings = payload.get('test_settings') or {}

    test_id = payload.get('test_id')
    question_id = payload.get('question_id')
    question_type = payload.get('type')
    text = (payload.get('text') or '').strip()
    points = payload.get('points', 1)
    image_file = request.FILES.get('image') if request.FILES else None
    image_remove = str(payload.get('image_remove') or '').lower() in {'1', 'true', 'yes'}

    if not test_id or not text:
        return JsonResponse({'error': 'Укажите тест и текст вопроса.'}, status=400)
    test = (
        Test.objects.select_related('course')
        .filter(id=test_id, course__created_by=request.user)
        .first()
    )
    if not test:
        return JsonResponse({'error': 'Тест не найден.'}, status=404)

    settings_payload = test_settings or {}
    def _to_int(value, default):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    title_override = (settings_payload.get('title') or '').strip()
    due_date = parse_date(settings_payload.get('due_date')) if settings_payload.get('due_date') else None
    overrides = {
        'title': title_override or None,
        'description': settings_payload.get('description'),
        'passing_score': _to_int(settings_payload.get('passing_score'), test.passing_score),
        'attempts': _to_int(settings_payload.get('attempts'), test.attempts),
        'warning_threshold': _to_int(settings_payload.get('warning_threshold'), test.warning_threshold),
        'success_threshold': _to_int(settings_payload.get('success_threshold'), test.success_threshold),
        'retake_requires_new_attempt': settings_payload.get('retake_requires_new_attempt'),
        'due_date': due_date,
        'is_published': settings_payload.get('is_published'),
    }

    versioned = False
    question_map = None
    if TestResult.objects.filter(test=test).exists():
        versioned = True
        if not overrides.get('title') or overrides.get('title') == test.title:
            base = test.title
            idx = 2
            while Test.objects.filter(course=test.course, title=f"{base} (версия {idx})").exists():
                idx += 1
            overrides['title'] = f"{base} (версия {idx})"
        test, question_map = _clone_test_with_questions(
            test,
            overrides=overrides,
            created_by=request.user,
            return_map=True,
        )
        if question_id and question_map:
            try:
                mapped = question_map.get(int(question_id))
            except (TypeError, ValueError):
                mapped = None
            if mapped:
                question_id = mapped
    else:
        update_fields = {}
        for field in (
            'title',
            'description',
            'passing_score',
            'attempts',
            'warning_threshold',
            'success_threshold',
            'retake_requires_new_attempt',
            'due_date',
            'is_published',
        ):
            if field in overrides and overrides[field] is not None:
                if field == 'title' and not str(overrides[field]).strip():
                    continue
                update_fields[field] = overrides[field]
        if update_fields:
            Test.objects.filter(pk=test.pk).update(**update_fields)
            test.refresh_from_db()

    allowed_types = {
        QUESTION_TYPE_SINGLE,
        QUESTION_TYPE_MULTI,
        QUESTION_TYPE_MATCHING,
        QUESTION_TYPE_ORDERING,
        QUESTION_TYPE_SHORT,
        QUESTION_TYPE_LONG,
    }
    if question_type not in allowed_types:
        return JsonResponse({'error': 'Этот тип вопроса пока не поддерживается.'}, status=400)

    try:
        points_value = int(points)
    except (TypeError, ValueError):
        points_value = 1

    if question_type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
        if not isinstance(answers, list) or len(answers) < 2:
            return JsonResponse({'error': 'Добавьте минимум два варианта ответа.'}, status=400)
        normalized = []
        correct_count = 0
        for item in answers:
            if not isinstance(item, dict):
                continue
            text_value = (item.get('text') or '').strip()
            if not text_value:
                continue
            is_correct = bool(item.get('is_correct'))
            if is_correct:
                correct_count += 1
            normalized.append({'text': text_value, 'is_correct': is_correct})
        if len(normalized) < 2:
            return JsonResponse({'error': 'Добавьте минимум два варианта ответа.'}, status=400)
        if question_type == QUESTION_TYPE_SINGLE and correct_count != 1:
            return JsonResponse({'error': 'Для одного правильного ответа выберите ровно один вариант.'}, status=400)
        if question_type == QUESTION_TYPE_MULTI and correct_count < 1:
            return JsonResponse({'error': 'Отметьте хотя бы один правильный ответ.'}, status=400)
        normalized_pairs = []
        normalized_ordering = []
    elif question_type == QUESTION_TYPE_MATCHING:
        if not isinstance(pairs, list) or len(pairs) < 2:
            return JsonResponse({'error': 'Добавьте минимум две пары.'}, status=400)
        normalized_pairs = []
        for item in pairs:
            if not isinstance(item, dict):
                continue
            left_text = (item.get('left_text') or '').strip()
            right_text = (item.get('right_text') or '').strip()
            if not left_text or not right_text:
                continue
            normalized_pairs.append({'left_text': left_text, 'right_text': right_text})
        if len(normalized_pairs) < 2:
            return JsonResponse({'error': 'Добавьте минимум две пары.'}, status=400)
        normalized = []
        normalized_ordering = []
    elif question_type == QUESTION_TYPE_ORDERING:
        if not isinstance(ordering, list) or len(ordering) < 2:
            return JsonResponse({'error': 'Добавьте минимум два элемента.'}, status=400)
        normalized_ordering = []
        for item in ordering:
            if not isinstance(item, dict):
                continue
            text_value = (item.get('text') or '').strip()
            if not text_value:
                continue
            normalized_ordering.append({'text': text_value})
        if len(normalized_ordering) < 2:
            return JsonResponse({'error': 'Добавьте минимум два элемента.'}, status=400)
        normalized = []
        normalized_pairs = []
    else:
        normalized = []
        normalized_pairs = []
        normalized_ordering = []

    def _apply_question_payload(target):
        target.type = question_type
        target.text = text
        target.points = points_value
        target.save()
        target.answers.all().delete()
        target.matching_pairs.all().delete()
        target.ordering_items.all().delete()
        if normalized:
            Answer.objects.bulk_create(
                [
                    Answer(question=target, text=item['text'], is_correct=item['is_correct'])
                    for item in normalized
                ]
            )
        if normalized_pairs:
            MatchingPair.objects.bulk_create(
                [
                    MatchingPair(
                        question=target,
                        left_text=item['left_text'],
                        right_text=item['right_text'],
                        order=index,
                    )
                    for index, item in enumerate(normalized_pairs)
                ]
            )
        if normalized_ordering:
            OrderingItem.objects.bulk_create(
                [
                    OrderingItem(
                        question=target,
                        text=item['text'],
                        position=index + 1,
                    )
                    for index, item in enumerate(normalized_ordering)
                ]
            )
        if image_remove:
            if target.image:
                target.image.delete(save=False)
            target.image = None
            target.save(update_fields=['image'])
        elif image_file:
            target.image = image_file
            target.save(update_fields=['image'])

    if question_id:
        question = (
            Question.objects.select_related('test')
            .filter(id=question_id, test=test)
            .first()
        )
        if not question:
            return JsonResponse({'error': 'Вопрос не найден.'}, status=404)
        _apply_question_payload(question)
    else:
        question = Question.objects.create(
            test=test,
            type=question_type,
            text=text,
            points=points_value,
        )
        if normalized:
            Answer.objects.bulk_create(
                [
                    Answer(question=question, text=item['text'], is_correct=item['is_correct'])
                    for item in normalized
                ]
            )
        if normalized_pairs:
            MatchingPair.objects.bulk_create(
                [
                    MatchingPair(
                        question=question,
                        left_text=item['left_text'],
                        right_text=item['right_text'],
                        order=index,
                    )
                    for index, item in enumerate(normalized_pairs)
                ]
            )
        if normalized_ordering:
            OrderingItem.objects.bulk_create(
                [
                    OrderingItem(
                        question=question,
                        text=item['text'],
                        position=index + 1,
                    )
                    for index, item in enumerate(normalized_ordering)
                ]
            )
        if image_file:
            question.image = image_file
            question.save(update_fields=['image'])
        if image_remove:
            if question.image:
                question.image.delete(save=False)
            question.image = None
            question.save(update_fields=['image'])
    return JsonResponse(
        {
            'id': question.id,
            'test_id': test.id,
            'test_title': test.title,
            'versioned': versioned,
            'question_map': question_map,
            'image_url': question.image.url if question.image else '',
        }
    )


@login_required
@role_required(['MANAGER'])
def manager_material_list(request):
    courses = Course.objects.filter(created_by=request.user).order_by('title')
    materials = (
        CourseMaterial.objects.filter(course__in=courses)
        .select_related('course', 'test')
        .order_by('course_id', 'order', 'id')
    )
    selected_course = request.GET.get('course')
    if selected_course:
        materials = materials.filter(course_id=selected_course)
    return render(
        request,
        'manager/material_list.html',
        {
            'materials': materials,
            'courses': courses,
            'selected_course': selected_course,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_material_create(request):
    courses_qs = Course.objects.filter(created_by=request.user).order_by('title')
    course_id = request.GET.get('course')
    test_id = request.GET.get('test')
    course = courses_qs.filter(id=course_id).first() if course_id else None
    if request.method == 'POST':
        course_id = request.POST.get('course')
        course = courses_qs.filter(id=course_id).first() if course_id else None
    tests_qs = Test.objects.filter(course=course) if course else Test.objects.filter(course__in=courses_qs)

    form = CourseMaterialForm(
        request.POST or None,
        request.FILES or None,
        course_qs=courses_qs,
        tests_qs=tests_qs,
    )
    if course and not form.is_bound:
        form.fields['course'].initial = course
    if test_id and not form.is_bound:
        form.fields['test'].initial = test_id

    if request.method == 'POST' and form.is_valid():
        files = form.cleaned_data.get('files') or []
        if files:
            metadata = {
                'course': form.cleaned_data['course'],
                'test': form.cleaned_data['test'],
                'title': form.cleaned_data['title'],
                'material_type': form.cleaned_data['material_type'],
                'content': form.cleaned_data['content'],
                'url': form.cleaned_data['url'],
                'accent_color': form.cleaned_data['accent_color'],
                'order': form.cleaned_data['order'],
                'is_required': form.cleaned_data['is_required'],
            }
            created_materials = []
            for uploaded_file in files:
                CourseMaterial.objects.create(
                    **{
                        **metadata,
                        'title': metadata['title'] or os.path.basename(uploaded_file.name),
                        'file': uploaded_file,
                        'image': None,
                    }
                )
            return redirect('manager_course_detail', course_id=metadata['course'].id)
        material = form.save()
        return redirect('manager_course_detail', course_id=material.course_id)

    return render(
        request,
        'manager/material_form.html',
        {
            'form': form,
            'is_edit': False,
            'course': course,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_material_edit(request, material_id):
    material = get_object_or_404(
        CourseMaterial.objects.select_related('course'),
        id=material_id,
        course__created_by=request.user,
    )
    courses_qs = Course.objects.filter(created_by=request.user).order_by('title')
    tests_qs = Test.objects.filter(course=material.course)
    form = CourseMaterialForm(
        request.POST or None,
        request.FILES or None,
        instance=material,
        course_qs=courses_qs,
        tests_qs=tests_qs,
    )
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('manager_course_detail', course_id=material.course_id)

    return render(
        request,
        'manager/material_form.html',
        {
            'form': form,
            'is_edit': True,
            'course': material.course,
            'material': material,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_material_delete(request, material_id):
    if request.method != 'POST':
        raise PermissionDenied
    material = get_object_or_404(
        CourseMaterial,
        id=material_id,
        course__created_by=request.user,
    )
    course_id = material.course_id
    material.delete()
    return redirect('manager_course_detail', course_id=course_id)


@login_required
@role_required(['MANAGER'])
def manager_students_list(request):
    students = (
        User.objects.filter(role='EMPLOYEE')
        .select_related('specialty', 'position')
        .order_by('last_name', 'first_name')
    )
    return render(request, 'manager/students_list.html', {'students': students})


@login_required
@role_required(['MANAGER'])
def manager_student_detail(request, student_id):
    student = get_object_or_404(
        User.objects.select_related('specialty', 'position'),
        id=student_id,
        role='EMPLOYEE',
    )
    if request.method == 'POST':
        raise PermissionDenied

    form = ManagerStudentForm(instance=student)
    for field in form.fields.values():
        field.disabled = True
    return render(
        request,
        'manager/student_detail.html',
        {
            'student': student,
            'form': form,
            'read_only': True,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_drafts(request):
    tasks = (
        Task.objects.filter(created_by=request.user, is_published=False)
        .select_related('course')
        .order_by('-created_at')
    )
    tests = (
        Test.objects.filter(created_by=request.user, is_published=False)
        .select_related('course')
        .order_by('-created_at')
    )
    return render(request, 'manager/drafts.html', {'tasks': tasks, 'tests': tests})


@login_required
@role_required(['MANAGER'])
def manager_task_publish(request, task_id):
    if request.method != 'POST':
        return redirect('manager_drafts')
    task = get_object_or_404(Task, id=task_id, created_by=request.user)
    if not task.is_published:
        task.is_published = True
        task.save(update_fields=['is_published'])
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('manager_drafts')


@login_required
@role_required(['MANAGER'])
def manager_test_publish(request, test_id):
    if request.method != 'POST':
        return redirect('manager_drafts')
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if not test.is_published:
        test.is_published = True
        test.save(update_fields=['is_published'])
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('manager_drafts')


@login_required
@role_required(['MANAGER'])
def manager_test_unpublish(request, test_id):
    if request.method != 'POST':
        return redirect('manager_drafts')
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if test.is_published:
        test.is_published = False
        test.save(update_fields=['is_published'])
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('manager_course_detail', course_id=test.course_id)


@login_required
@role_required(['MANAGER'])
def manager_test_delete(request, test_id):
    if request.method != 'POST':
        return redirect('manager_drafts')
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    course_id = test.course_id
    test.delete()
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    if course_id:
        return redirect('manager_course_detail', course_id=course_id)
    return redirect('manager_drafts')
@login_required
@role_required(['MANAGER'])
def manager_course_create(request):
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES)
        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()
            return redirect('manager_course_detail', course_id=course.id)
    else:
        form = CourseForm()
    return render(request, 'manager/course_create.html', {'form': form, 'is_edit': False})


@login_required
@role_required(['MANAGER'])
def manager_course_edit(request, course_id):
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            return redirect('manager_course_detail', course_id=course.id)
    else:
        form = CourseForm(instance=course)
    return render(
        request,
        'manager/course_create.html',
        {'form': form, 'is_edit': True, 'course': course},
    )


@login_required
@role_required(['MANAGER'])
def manager_course_delete(request, course_id):
    if request.method != 'POST':
        return redirect('manager_course_list')
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    course.delete()
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('manager_course_list')


@login_required
@role_required(['MANAGER'])
def manager_course_assign(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    users_qs = User.objects.filter(role='EMPLOYEE')
    if course.specialty_id:
        users_qs = users_qs.filter(specialty=course.specialty)
    users_count = users_qs.count()

    if request.method == 'POST':
        form = CourseAssignForm(request.POST, users_qs=users_qs)
        if form.is_valid():
            due_date = form.cleaned_data.get('due_date')
            if form.cleaned_data.get('assign_all'):
                selected_users = users_qs
            else:
                selected_users = form.cleaned_data['users']
            for user in selected_users:
                Enrollment.objects.get_or_create(
                    user=user,
                    course=course,
                    defaults={
                        'assigned_by': request.user,
                        'due_date': due_date,
                        'status': 'ASSIGNED',
                        'progress': 0,
                        'completed': False,
                    },
                )
            return redirect('manager_course_detail', course_id=course.id)
    else:
        form = CourseAssignForm(users_qs=users_qs)

    return render(
        request,
        'manager/course_assign.html',
        {'form': form, 'course': course, 'users_count': users_count},
    )


@login_required
@role_required(['MANAGER'])
def manager_test_create(request):
    course_id = request.GET.get('course')
    course = None
    if course_id:
        course = get_object_or_404(Course, id=course_id, created_by=request.user)
    if request.method == 'POST':
        form = TestForm(request.POST)
        if form.is_valid():
            test = form.save(commit=False)
            test.created_by = request.user
            if course:
                test.course = course
            test.save()
            return redirect('manager_test_questions', test_id=test.id)
    else:
        form = TestForm(initial={'course': course} if course else None)
        if course:
            form.fields['course'].disabled = True
    return render(request, 'manager/test_create.html', {'form': form, 'is_edit': False})


@login_required
@role_required(['MANAGER'])
def manager_test_edit(request, test_id):
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if _test_is_assigned(test):
        raise PermissionDenied
    if request.method == 'POST':
        form = TestForm(request.POST, instance=test)
        if form.is_valid():
            form.save()
            return redirect('manager_course_detail', course_id=test.course_id)
    else:
        form = TestForm(instance=test)
    return render(
        request,
        'manager/test_create.html',
        {'form': form, 'is_edit': True, 'test': test},
    )


@login_required
@role_required(['MANAGER'])
def manager_test_questions(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    can_edit = not _test_is_assigned(test)
    if request.method == 'POST':
        if not can_edit:
            raise PermissionDenied
        q_form = QuestionForm(request.POST, request.FILES)
        answer_formset = AnswerFormSet(request.POST, prefix='answers')
        matching_formset = MatchingPairFormSet(request.POST, prefix='matching')
        ordering_formset = OrderingItemFormSet(request.POST, prefix='ordering')
        formset = None
        formsets_valid = True
        if q_form.is_valid():
            question_type = q_form.cleaned_data['type']
            if question_type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
                formset = answer_formset
                formsets_valid = formset.is_valid()
                if formsets_valid:
                    correct_count = 0
                    for form in formset:
                        if not getattr(form, 'cleaned_data', None):
                            continue
                        if form.cleaned_data.get('DELETE'):
                            continue
                        if form.cleaned_data.get('is_correct'):
                            correct_count += 1
                    if question_type == QUESTION_TYPE_SINGLE and correct_count != 1:
                        formset._non_form_errors = formset.error_class(
                            ['Для вопроса с одним вариантом нужен ровно один правильный ответ.']
                        )
                        formsets_valid = False
                    if question_type == QUESTION_TYPE_MULTI and correct_count < 1:
                        formset._non_form_errors = formset.error_class(
                            ['Для вопроса с несколькими вариантами нужен хотя бы один правильный ответ.']
                        )
                        formsets_valid = False
            elif question_type == QUESTION_TYPE_MATCHING:
                formset = matching_formset
                formsets_valid = formset.is_valid()
            elif question_type == QUESTION_TYPE_ORDERING:
                formset = ordering_formset
                formsets_valid = formset.is_valid()
            elif question_type in {QUESTION_TYPE_SHORT, QUESTION_TYPE_LONG}:
                formset = None
                formsets_valid = True
        if q_form.is_valid() and formsets_valid:
            question = q_form.save(commit=False)
            question.test = test
            question.save()
            if formset is not None:
                formset.instance = question
                if question_type == QUESTION_TYPE_MATCHING:
                    pairs = formset.save(commit=False)
                    order_value = 1
                    for pair in pairs:
                        pair.order = order_value
                        pair.save()
                        order_value += 1
                else:
                    formset.save()
            return redirect('manager_test_questions', test_id=test.id)
    else:
        q_form = QuestionForm()
        answer_formset = AnswerFormSet(prefix='answers')
        matching_formset = MatchingPairFormSet(prefix='matching')
        ordering_formset = OrderingItemFormSet(prefix='ordering')

    questions = Question.objects.filter(test=test).prefetch_related(
        'answers',
        'matching_pairs',
        'ordering_items',
    )
    return render(
        request,
        'manager/test_questions.html',
        {
            'test': test,
            'q_form': q_form,
            'answer_formset': answer_formset,
            'matching_formset': matching_formset,
            'ordering_formset': ordering_formset,
            'questions': questions,
            'can_edit': can_edit,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_test_results(request, test_id):
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    results = (
        TestResult.objects.filter(test=test)
        .select_related('user')
        .order_by('-completed_at', '-id')
    )
    return render(
        request,
        'manager/test_results.html',
        {'test': test, 'results': results},
    )


@login_required
@role_required(['MANAGER'])
def manager_test_review(request, result_id):
    result = get_object_or_404(
        TestResult.objects.select_related('test', 'user'),
        id=result_id,
    )
    if result.test.created_by_id != request.user.id:
        raise PermissionDenied

    feedback = Feedback.objects.filter(test_result=result).order_by('-created_at').first()
    questions = (
        Question.objects.filter(test=result.test)
        .prefetch_related('answers', 'matching_pairs', 'ordering_items')
        .order_by('id')
    )
    selections = (
        TestAnswer.objects.filter(test_result=result)
        .select_related('answer', 'question')
    )
    answers_by_question = {}
    for sel in selections:
        answers_by_question.setdefault(sel.question_id, []).append(sel)

    result_questions = []
    for question in questions:
        question_type = question.type
        item = {'question': question, 'type': question_type}
        if question_type in {QUESTION_TYPE_SINGLE, QUESTION_TYPE_MULTI}:
            answers = list(question.answers.all())
            correct_ids = {answer.id for answer in answers if answer.is_correct}
            selected_ids = {
                sel.answer_id
                for sel in answers_by_question.get(question.id, [])
                if sel.answer_id
            }
            item.update(
                {
                    'answers': answers,
                    'correct_ids': correct_ids,
                    'selected_ids': selected_ids,
                }
            )
        elif question_type == QUESTION_TYPE_MATCHING:
            pairs = list(question.matching_pairs.all().order_by('order', 'id'))
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_data),
                None,
            )
            match_map = {}
            if selection and selection.answer_data:
                for match in selection.answer_data.get('matches', []):
                    left_id = match.get('left_id')
                    right_id = match.get('right_id')
                    if left_id is not None and right_id is not None:
                        match_map[left_id] = right_id
            right_by_id = {pair.id: pair.right_text for pair in pairs}
            matching_rows = []
            for pair in pairs:
                selected_id = match_map.get(pair.id)
                matching_rows.append(
                    {
                        'left_text': pair.left_text,
                        'right_text': pair.right_text,
                        'selected_right_text': right_by_id.get(selected_id),
                        'selected_right_id': selected_id,
                        'right_id': pair.id,
                    }
                )
            item.update({'matching_rows': matching_rows})
        elif question_type == QUESTION_TYPE_ORDERING:
            ordering_items = list(question.ordering_items.all().order_by('position', 'id'))
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_data),
                None,
            )
            order_list = []
            positions_map = {}
            if selection and selection.answer_data:
                order_list = selection.answer_data.get('order') or []
                for row in selection.answer_data.get('positions') or []:
                    item_id = row.get('id')
                    position = row.get('position')
                    if item_id is not None:
                        positions_map[item_id] = position
            if not positions_map and order_list:
                for idx, item_id in enumerate(order_list, start=1):
                    positions_map[item_id] = idx
            ordering_rows = [
                {
                    'id': item.id,
                    'text': item.text,
                    'position': item.position,
                    'user_position': positions_map.get(item.id),
                }
                for item in ordering_items
            ]
            item.update({'ordering_rows': ordering_rows})
        elif question_type in {QUESTION_TYPE_SHORT, QUESTION_TYPE_LONG}:
            selection = next(
                (sel for sel in answers_by_question.get(question.id, []) if sel.answer_text is not None),
                None,
            )
            item.update(
                {
                    'answer_text': selection.answer_text if selection else '',
                    'answer_file': selection.attachment if selection else None,
                }
            )
        result_questions.append(item)

    error = None
    if request.method == 'POST':
        status_value = request.POST.get('status')
        score_raw = request.POST.get('score', '').strip()
        comment = request.POST.get('comment', '').strip()

        allowed_statuses = {
            TestResult.Status.PASSED,
            TestResult.Status.FAILED,
            TestResult.Status.RETURNED,
        }
        if status_value not in allowed_statuses:
            error = 'Некорректный статус.'
        else:
            score = None
            passed = None
            if status_value in [TestResult.Status.PASSED, TestResult.Status.FAILED]:
                if score_raw == '':
                    error = 'Укажите процент.'
                else:
                    try:
                        score = int(score_raw)
                    except ValueError:
                        error = 'Процент должен быть числом.'
                    else:
                        if score < 0:
                            error = 'Процент должен быть не меньше 0.'
                        elif score > 100:
                            error = 'Процент не может быть больше 100.'
                passed = status_value == TestResult.Status.PASSED

            if not error:
                result.status = status_value
                result.score = score
                result.passed = passed
                if status_value == TestResult.Status.RETURNED:
                    result.retake_accepted = False
                    result.save(update_fields=['status', 'score', 'passed', 'retake_accepted'])
                else:
                    result.save(update_fields=['status', 'score', 'passed'])

                if comment or score is not None:
                    Feedback.objects.update_or_create(
                        test_result=result,
                        defaults={
                            'manager': request.user,
                            'comment': comment,
                            'rating': score,
                        },
                    )
                else:
                    Feedback.objects.filter(test_result=result).delete()

                return redirect('manager_test_results', test_id=result.test_id)

    return render(
        request,
        'manager/test_review.html',
        {
            'result': result,
            'test': result.test,
            'feedback': feedback,
            'result_questions': result_questions,
            'error': error,
        },
    )


@login_required
@role_required(['MANAGER'])
def manager_task_create(request):
    course_id = request.GET.get('course')
    course = None
    if course_id:
        course = get_object_or_404(Course, id=course_id, created_by=request.user)
    if request.method == 'POST':
        form = TaskForm(request.POST, request.FILES)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            if course:
                task.course = course
            task.save()
            return redirect('manager_task_assign', task_id=task.id)
    else:
        form = TaskForm(initial={'course': course} if course else None)
        if course:
            form.fields['course'].disabled = True
    return render(request, 'manager/task_create.html', {'form': form, 'is_edit': False})


@login_required
@role_required(['MANAGER'])
def manager_task_edit(request, task_id):
    task = get_object_or_404(Task, id=task_id, created_by=request.user)
    if _task_has_submissions(task):
        raise PermissionDenied
    if request.method == 'POST':
        form = TaskForm(request.POST, request.FILES, instance=task)
        if form.is_valid():
            form.save()
            return redirect('manager_task_list')
    else:
        form = TaskForm(instance=task)
    return render(
        request,
        'manager/task_create.html',
        {'form': form, 'is_edit': True, 'task': task},
    )


@login_required
@role_required(['MANAGER'])
def manager_task_delete(request, task_id):
    if request.method != 'POST':
        return redirect('manager_task_list')
    task = get_object_or_404(Task, id=task_id, created_by=request.user)
    task.delete()
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('manager_task_list')


@login_required
@role_required(['MANAGER'])
def manager_task_assign(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    users_qs = (
        User.objects.filter(role='EMPLOYEE')
        .select_related('position')
        .order_by('last_name', 'first_name', 'email')
    )
    if task.course_id and task.course.specialty_id:
        users_qs = users_qs.filter(specialty=task.course.specialty)
    users_count = users_qs.count()
    assigned_user_ids = set(
        TaskAssignment.objects.filter(task=task).values_list('user_id', flat=True)
    )

    if request.method == 'POST':
        revoke_user_id = (request.POST.get('revoke_user') or '').strip()
        if revoke_user_id.isdigit():
            TaskAssignment.objects.filter(task=task, user_id=int(revoke_user_id)).delete()
            return redirect('manager_task_assign', task_id=task.id)
        form = TaskAssignForm(request.POST, users_qs=users_qs)
        if form.is_valid():
            due_date = form.cleaned_data.get('due_date') or task.due_date
            priority = form.cleaned_data.get('priority')
            users = (
                list(users_qs)
                if form.cleaned_data.get('assign_all')
                else form.cleaned_data.get('users')
            )
            for user in users:
                TaskAssignment.objects.get_or_create(
                    task=task,
                    user=user,
                    defaults={
                        'assigned_by': request.user,
                        'due_date': due_date,
                        'priority': priority,
                        'status': 'ASSIGNED',
                    },
                )
            return redirect('manager_dashboard')
    else:
        form = TaskAssignForm(users_qs=users_qs, initial={'due_date': task.due_date})

    return render(
        request,
        'manager/task_assign.html',
        {
            'form': form,
            'task': task,
            'users_count': users_count,
            'users': list(users_qs),
            'assigned_user_ids': assigned_user_ids,
        },
    )


@login_required
@role_required(['ANALYST'])
def analyst_dashboard(request):
    filters = _analytics_filters_from_request(request)
    if request.method == 'POST':
        action = request.POST.get('action')
        filters = _analytics_filters_from_payload(request.POST, base=filters)
        title = (request.POST.get('title') or '').strip()
        if action == 'save_dashboard' and title:
            Dashboard.objects.create(owner=request.user, title=title, config=filters)
        elif action == 'save_report' and title:
            report_type = (request.POST.get('report_type') or Report.ReportType.PROGRESS)
            Report.objects.create(
                owner=request.user,
                title=title,
                report_type=report_type,
                filters=filters,
            )
        qs = urlencode({k: v for k, v in filters.items() if v})
        return redirect(f"{reverse('analyst_dashboard')}?{qs}" if qs else reverse('analyst_dashboard'))

    data = _build_analytics_data(filters)
    dashboards = Dashboard.objects.filter(Q(owner=request.user) | Q(is_shared=True)).order_by('-updated_at')
    reports = Report.objects.filter(owner=request.user).order_by('-updated_at')
    departments = (
        User.objects.filter(role='EMPLOYEE')
        .exclude(department__isnull=True)
        .exclude(department='')
        .values_list('department', flat=True)
        .distinct()
        .order_by('department')
    )
    specialties = Specialty.objects.filter(is_active=True).order_by('name')
    positions = Position.objects.filter(is_active=True).order_by('name')
    courses = Course.objects.filter(is_active=True).order_by('title')

    return render(
        request,
        'users/analyst_dashboard.html',
        {
            'filters': filters,
            'summary': data['summary'],
            'employee_rows': data['employee_rows'],
            'department_rows': data['department_rows'],
            'monthly_stats': data['monthly_stats'],
              'monthly_series': data['monthly_series'],
              'dept_series': data['dept_series'],
              'status_series': data['status_series'],
              'tasks_series': data['tasks_series'],
              'competencies_rows': data['competencies_rows'],
            'recent_tests': data['recent_tests'],
            'recent_tasks': data['recent_tasks'],
            'dashboards': dashboards,
            'reports': reports,
            'departments': departments,
            'specialties': specialties,
            'positions': positions,
            'courses': courses,
        },
    )


@login_required
@role_required(['ANALYST'])
def analyst_export(request, export_format):
    export_format = (export_format or '').lower()
    filters = _analytics_filters_from_request(request)
    data = _build_analytics_data(filters)

    if export_format == 'csv':
        output = io.StringIO(newline='')
        # Excel in RU locales typically expects ';' as delimiter and UTF‑8 BOM.
        writer = csv.writer(output, delimiter=';')
        summary = data['summary']
        writer.writerow(['Сводка'])
        writer.writerow(['Сотрудники', summary['employees']])
        writer.writerow(['Курсы назначены', summary['courses_assigned']])
        writer.writerow(['Курсы завершены', summary['courses_completed']])
        writer.writerow(['Курсы в процессе', summary['courses_in_progress']])
        writer.writerow(['Курсы просрочены', summary['courses_overdue']])
        writer.writerow(['Тесты пройдены', summary['tests_taken']])
        writer.writerow(['Средний балл', summary['avg_score'] or 0])
        writer.writerow(['Процент прохождения', summary['pass_rate']])
        writer.writerow(['Задания отправлено', summary['tasks_submitted']])
        writer.writerow(['Задания приняты', summary['tasks_approved']])
        writer.writerow([])
        writer.writerow(['Агрегация по сотрудникам'])
        writer.writerow([
            'Сотрудник',
            'Отдел',
            'Специальность',
            'Должность',
            'Курсы назначены',
            'Курсы завершены',
            'Курсы в процессе',
            'Курсы просрочены',
            'Средний прогресс',
            'Тесты пройдены',
            'Средний балл',
            'Процент прохождения',
            'Задания отправлено',
            'Задания приняты',
        ])
        for row in data['employee_rows']:
            user = row['user']
            writer.writerow([
                f'{user.last_name} {user.first_name}'.strip() or user.email,
                user.department or '—',
                user.specialty or '—',
                user.position or '—',
                row['assigned'],
                row['completed'],
                row['in_progress'],
                row['overdue'],
                round(row['progress_avg'], 1) if row['progress_avg'] is not None else '—',
                row['tests_taken'],
                round(row['avg_score'], 1) if row['avg_score'] is not None else '—',
                row['pass_rate'],
                row['tasks_submitted'],
                row['tasks_approved'],
            ])
        writer.writerow([])
        writer.writerow(['Агрегация по подразделениям'])
        writer.writerow([
            'Подразделение',
            'Курсы назначены',
            'Курсы завершены',
            'Курсы в процессе',
            'Курсы просрочены',
            'Средний прогресс',
            'Тесты пройдены',
            'Средний балл',
            'Процент прохождения',
            'Задания отправлено',
            'Задания приняты',
        ])
        for row in data['department_rows']:
            writer.writerow([
                row['department'] or '—',
                row['assigned'],
                row['completed'],
                row['in_progress'],
                row['overdue'],
                round(row['progress_avg'], 1) if row['progress_avg'] is not None else '—',
                row['tests_taken'],
                round(row['avg_score'], 1) if row['avg_score'] is not None else '—',
                row['pass_rate'],
                row['tasks_submitted'],
                row['tasks_approved'],
            ])

        payload = output.getvalue().encode('utf-8-sig')
        response = HttpResponse(payload, content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="analytics-report.csv"'
        return response

    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            return HttpResponse('XLSX экспорт недоступен. Установите openpyxl.', status=500)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Сводка'

        bold = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical='top')

        summary = data['summary']
        summary_rows = [
            ('Сотрудники', summary['employees']),
            ('Курсы назначены', summary['courses_assigned']),
            ('Курсы завершены', summary['courses_completed']),
            ('Курсы в процессе', summary['courses_in_progress']),
            ('Курсы просрочены', summary['courses_overdue']),
            ('Тесты пройдены', summary['tests_taken']),
            ('Средний балл', float(summary['avg_score'] or 0)),
            ('Процент прохождения', f"{summary['pass_rate']}%"),
            ('Задания отправлено', summary['tasks_submitted']),
            ('Задания приняты', summary['tasks_approved']),
        ]
        ws_summary.append(['Показатель', 'Значение'])
        ws_summary['A1'].font = bold
        ws_summary['B1'].font = bold
        for k, v in summary_rows:
            ws_summary.append([k, v])
        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 18

        ws_emp = wb.create_sheet('Сотрудники')
        ws_emp.append([
            'Сотрудник',
            'Отдел',
            'Специальность',
            'Должность',
            'Курсы назначены',
            'Курсы завершены',
            'Курсы в процессе',
            'Курсы просрочены',
            'Средний прогресс',
            'Тесты пройдены',
            'Средний балл',
            'Процент прохождения',
            'Задания отправлено',
            'Задания приняты',
        ])
        for cell in ws_emp[1]:
            cell.font = bold
            cell.alignment = wrap
        for row in data['employee_rows']:
            user = row['user']
            ws_emp.append([
                f'{user.last_name} {user.first_name}'.strip() or user.email,
                user.department or '—',
                str(user.specialty or '—'),
                str(user.position or '—'),
                row['assigned'],
                row['completed'],
                row['in_progress'],
                row['overdue'],
                round(row['progress_avg'], 1) if row['progress_avg'] is not None else None,
                row['tests_taken'],
                round(row['avg_score'], 1) if row['avg_score'] is not None else None,
                f"{row['pass_rate']}%",
                row['tasks_submitted'],
                row['tasks_approved'],
            ])
        ws_emp.freeze_panes = 'A2'

        ws_dept = wb.create_sheet('Подразделения')
        ws_dept.append([
            'Подразделение',
            'Курсы назначены',
            'Курсы завершены',
            'Курсы в процессе',
            'Курсы просрочены',
            'Средний прогресс',
            'Тесты пройдены',
            'Средний балл',
            'Процент прохождения',
            'Задания отправлено',
            'Задания приняты',
        ])
        for cell in ws_dept[1]:
            cell.font = bold
            cell.alignment = wrap
        for row in data['department_rows']:
            ws_dept.append([
                row['department'] or '—',
                row['assigned'],
                row['completed'],
                row['in_progress'],
                row['overdue'],
                round(row['progress_avg'], 1) if row['progress_avg'] is not None else None,
                row['tests_taken'],
                round(row['avg_score'], 1) if row['avg_score'] is not None else None,
                f"{row['pass_rate']}%",
                row['tasks_submitted'],
                row['tasks_approved'],
            ])
        ws_dept.freeze_panes = 'A2'
        ws_dept.column_dimensions['A'].width = 28

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="analytics-report.xlsx"'
        return response

    if export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen import canvas
        except ImportError:
            return HttpResponse('PDF экспорт недоступен. Установите reportlab.', status=500)

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        font_dir = Path(settings.BASE_DIR) / 'static' / 'fonts'
        regular_font = font_dir / 'DejaVuSans.ttf'
        bold_font = font_dir / 'DejaVuSans-Bold.ttf'
        if regular_font.exists() and bold_font.exists():
            pdfmetrics.registerFont(TTFont('DejaVuSans', str(regular_font)))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', str(bold_font)))
            font_regular = 'DejaVuSans'
            font_bold = 'DejaVuSans-Bold'
        else:
            font_regular = 'Helvetica'
            font_bold = 'Helvetica-Bold'

        c.setFont(font_bold, 14)
        c.drawString(40, y, 'Отчёт по обучению')
        y -= 24
        c.setFont(font_regular, 10)
        summary = data['summary']
        lines = [
            f"Сотрудники: {summary['employees']}",
            f"Курсы назначены: {summary['courses_assigned']}",
            f"Курсы завершены: {summary['courses_completed']}",
            f"Курсы в процессе: {summary['courses_in_progress']}",
            f"Курсы просрочены: {summary['courses_overdue']}",
            f"Тесты пройдены: {summary['tests_taken']}",
            f"Средний балл: {round(summary['avg_score'], 1) if summary['avg_score'] is not None else '—'}",
            f"Процент прохождения: {summary['pass_rate']}%",
            f"Задания отправлено: {summary['tasks_submitted']}",
            f"Задания приняты: {summary['tasks_approved']}",
        ]
        for line in lines:
            c.drawString(40, y, line)
            y -= 14
        y -= 10
        c.setFont(font_bold, 11)
        c.drawString(40, y, 'Топ сотрудников')
        y -= 16
        c.setFont(font_regular, 9)
        for row in data['employee_rows'][:20]:
            user = row['user']
            label = f"{user.last_name} {user.first_name}".strip() or user.email
            line = f"{label} — тесты: {row['tests_taken']}, средний балл: {round(row['avg_score'], 1) if row['avg_score'] is not None else '—'}"
            c.drawString(40, y, line[:110])
            y -= 12
            if y < 60:
                c.showPage()
                y = height - 40
        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="analytics-report.pdf"'
        return response

    return HttpResponse('Неверный формат экспорта.', status=400)

def home_view(request):
    if request.user.is_authenticated:
        return redirect('/redirect/')
    return render(request, 'users/home.html')


def register_view(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            return render(
                request,
                'users/register.html',
                {
                    'form': UserRegisterForm(),
                    'submitted': True,
                    'submitted_email': user.email,
                },
            )
    else:
        form = UserRegisterForm()

    return render(request, 'users/register.html', {'form': form, 'submitted': False})

